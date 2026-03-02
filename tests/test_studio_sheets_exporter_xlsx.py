"""Tests for core/studio/sheets/exporter_xlsx.py — XLSX export via openpyxl."""

import openpyxl
import pytest

from core.schemas.studio_schema import SheetContentTree, SheetTab
from core.studio.sheets.exporter_xlsx import export_to_xlsx


def _make_tree(**kwargs) -> SheetContentTree:
    defaults = {
        "workbook_title": "Test Workbook",
        "tabs": [
            SheetTab(
                id="t1",
                name="Revenue",
                headers=["Month", "Amount"],
                rows=[["Jan", 100], ["Feb", 200]],
                formulas={"C2": "=B2*1.1"},
                column_widths=[120, 80],
            )
        ],
    }
    defaults.update(kwargs)
    return SheetContentTree(**defaults)


def test_xlsx_creates_file(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    assert path.exists()


def test_xlsx_creates_parent_dirs(tmp_path):
    path = tmp_path / "sub" / "dir" / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    assert path.exists()


def test_xlsx_opens_with_openpyxl(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    wb.close()


def test_xlsx_sheet_count_matches_tabs(tmp_path):
    tree = _make_tree(
        tabs=[
            SheetTab(id="t1", name="Sheet1", headers=["A"], rows=[[1]]),
            SheetTab(id="t2", name="Sheet2", headers=["B"], rows=[[2]]),
        ]
    )
    path = tmp_path / "output.xlsx"
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    assert len(wb.sheetnames) == 2
    wb.close()


def test_xlsx_sheet_names_match(tmp_path):
    tree = _make_tree(
        tabs=[
            SheetTab(id="t1", name="Revenue", headers=["A"], rows=[[1]]),
            SheetTab(id="t2", name="Costs", headers=["B"], rows=[[2]]),
        ]
    )
    path = tmp_path / "output.xlsx"
    export_to_xlsx(tree, path)
    wb = openpyxl.load_workbook(str(path))
    assert wb.sheetnames == ["Revenue", "Costs"]
    wb.close()


def test_xlsx_header_row_present(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, 3)]
    assert headers == ["Month", "Amount"]
    wb.close()


def test_xlsx_data_rows_present(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    row1 = [ws.cell(row=2, column=c).value for c in range(1, 3)]
    assert row1 == ["Jan", 100]
    wb.close()


def test_xlsx_formulas_written(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws["C2"].value == "=B2*1.1"
    wb.close()


def test_xlsx_column_widths_applied(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    # Column A should have a width based on 120/7
    assert ws.column_dimensions["A"].width > 0
    wb.close()


def test_xlsx_frozen_pane(tmp_path):
    path = tmp_path / "output.xlsx"
    export_to_xlsx(_make_tree(), path)
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    assert ws.freeze_panes == "A2"
    wb.close()
