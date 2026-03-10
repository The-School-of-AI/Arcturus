"""Tests for core/studio/sheets/validator.py — XLSX, CSV, and CSV-ZIP validators."""

import csv
import zipfile

import openpyxl

from core.studio.sheets.validator import validate_csv, validate_csv_zip, validate_xlsx


# === XLSX validation ===


def _create_valid_xlsx(path, sheets=None):
    """Helper to create a valid XLSX file."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, headers, rows, formulas in (sheets or [("Sheet1", ["A", "B"], [[1, 2]], {})]):
        ws = wb.create_sheet(title=name)
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        for addr, formula in formulas.items():
            ws[addr] = formula
    wb.save(str(path))


def test_validate_xlsx_success(tmp_path):
    path = tmp_path / "valid.xlsx"
    _create_valid_xlsx(path, [("Data", ["X", "Y"], [[1, 2]], {"C2": "=A2+B2"})])
    result = validate_xlsx(path, expected_sheet_names=["Data"], expected_formula_cells=1)
    assert result["valid"] is True
    assert result["sheet_count"] == 1
    assert "Data" in result["sheet_names"]
    assert result["formula_cell_count"] == 1


def test_validate_xlsx_corrupt_file(tmp_path):
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"not a valid xlsx")
    result = validate_xlsx(path)
    assert result["valid"] is False
    assert len(result["errors"]) > 0


def test_validate_xlsx_missing_sheets(tmp_path):
    path = tmp_path / "valid.xlsx"
    _create_valid_xlsx(path, [("Data", ["X"], [[1]], {})])
    result = validate_xlsx(path, expected_sheet_names=["Data", "Missing"])
    assert result["valid"] is False
    assert any("Missing" in e for e in result["errors"])


def test_validate_xlsx_formula_count(tmp_path):
    path = tmp_path / "valid.xlsx"
    _create_valid_xlsx(path, [("Data", ["X"], [[1]], {"B2": "=A2*2"})])
    result = validate_xlsx(path, expected_formula_cells=5)
    assert result["valid"] is False


def test_validate_xlsx_result_contract(tmp_path):
    path = tmp_path / "valid.xlsx"
    _create_valid_xlsx(path)
    result = validate_xlsx(path)
    assert "valid" in result
    assert "format" in result
    assert "errors" in result
    assert "warnings" in result
    assert "sheet_count" in result
    assert "sheet_names" in result
    assert "chart_count" in result
    assert "conditional_format_count" in result
    assert "styled_header_sheet_count" in result
    assert "quality_score" in result
    assert result["format"] == "xlsx"


def test_validate_xlsx_quality_score_improves_with_styling_and_chart(tmp_path):
    plain_path = tmp_path / "plain.xlsx"
    rich_path = tmp_path / "rich.xlsx"

    _create_valid_xlsx(
        plain_path,
        [("Data", ["Month", "Revenue"], [["Jan", 100], ["Feb", 200], ["Mar", 300]], {})],
    )
    _create_valid_xlsx(
        rich_path,
        [("Data", ["Month", "Revenue"], [["Jan", 100], ["Feb", 200], ["Mar", 300]], {})],
    )

    wb = openpyxl.load_workbook(str(rich_path))
    ws = wb["Data"]

    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Font, PatternFill

    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws["B1"].font = Font(bold=True)
    ws["B1"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=4))
    ws.add_chart(chart, "D2")

    ws.conditional_formatting.add(
        "B2:B4",
        ColorScaleRule(
            start_type="min",
            start_color="FEE2E2",
            mid_type="percentile",
            mid_value=50,
            mid_color="FEF3C7",
            end_type="max",
            end_color="DCFCE7",
        ),
    )
    wb.save(str(rich_path))
    wb.close()

    plain_result = validate_xlsx(plain_path)
    rich_result = validate_xlsx(rich_path)

    assert rich_result["quality_score"] > plain_result["quality_score"]
    assert rich_result["chart_count"] >= 1
    assert rich_result["conditional_format_count"] >= 1


# === CSV validation ===


def _create_valid_csv(path, headers=None, rows=None):
    headers = headers or ["A", "B"]
    rows = rows or [["1", "2"], ["3", "4"]]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


def test_validate_csv_success(tmp_path):
    path = tmp_path / "valid.csv"
    _create_valid_csv(path)
    result = validate_csv(path)
    assert result["valid"] is True
    assert result["row_count"] == 2
    assert result["column_count"] == 2


def test_validate_csv_empty_file(tmp_path):
    path = tmp_path / "empty.csv"
    path.write_text("")
    result = validate_csv(path)
    assert result["valid"] is False
    assert any("empty" in e.lower() for e in result["errors"])


def test_validate_csv_min_rows(tmp_path):
    path = tmp_path / "short.csv"
    _create_valid_csv(path, rows=[["1", "2"]])
    result = validate_csv(path, min_rows=5)
    assert result["valid"] is False
    assert any("5" in e for e in result["errors"])


def test_validate_csv_result_contract(tmp_path):
    path = tmp_path / "valid.csv"
    _create_valid_csv(path)
    result = validate_csv(path)
    assert "valid" in result
    assert "format" in result
    assert "errors" in result
    assert "warnings" in result
    assert "row_count" in result
    assert "column_count" in result
    assert result["format"] == "csv"


def test_validators_never_throw(tmp_path):
    # XLSX: nonexistent file
    result = validate_xlsx(tmp_path / "nope.xlsx")
    assert result["valid"] is False

    # CSV: nonexistent file
    result = validate_csv(tmp_path / "nope.csv")
    assert result["valid"] is False


# === CSV-ZIP validation ===


def _create_csv_zip(path, tabs):
    """Helper: create a ZIP with one CSV per tab.

    tabs is a list of (filename, headers, rows) tuples.
    """
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for filename, headers, rows in tabs:
            import io
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            zf.writestr(filename, buf.getvalue().encode("utf-8"))


class TestValidateCsvZip:
    def test_valid_zip(self, tmp_path):
        path = tmp_path / "valid.zip"
        _create_csv_zip(path, [
            ("Sheet1.csv", ["A", "B"], [["1", "2"]]),
            ("Sheet2.csv", ["X", "Y"], [["3", "4"]]),
        ])
        result = validate_csv_zip(path)
        assert result["valid"] is True
        assert result["tab_count"] == 2
        assert result["format"] == "csv_zip"
        assert len(result["per_tab"]) == 2

    def test_not_a_zip(self, tmp_path):
        path = tmp_path / "bad.zip"
        path.write_bytes(b"this is not a zip file")
        result = validate_csv_zip(path)
        assert result["valid"] is False
        assert any("not a valid ZIP" in e for e in result["errors"])

    def test_empty_zip(self, tmp_path):
        path = tmp_path / "empty.zip"
        with zipfile.ZipFile(path, "w") as zf:
            pass  # no members
        result = validate_csv_zip(path)
        assert result["valid"] is False
        assert any("no .csv files" in e for e in result["errors"])

    def test_wrong_tab_count(self, tmp_path):
        path = tmp_path / "valid.zip"
        _create_csv_zip(path, [("A.csv", ["X"], [["1"]])])
        result = validate_csv_zip(path, expected_tab_names=["A", "B"])
        assert result["valid"] is False
        assert any("Expected 2" in e for e in result["errors"])

    def test_per_tab_results_present(self, tmp_path):
        path = tmp_path / "valid.zip"
        _create_csv_zip(path, [
            ("Tab.csv", ["Col"], [["val"]]),
        ])
        result = validate_csv_zip(path)
        assert len(result["per_tab"]) == 1
        assert result["per_tab"][0]["filename"] == "Tab.csv"
        assert result["per_tab"][0]["valid"] is True

    def test_result_contract_keys(self, tmp_path):
        path = tmp_path / "valid.zip"
        _create_csv_zip(path, [("S.csv", ["A"], [["1"]])])
        result = validate_csv_zip(path)
        for key in ("valid", "format", "errors", "warnings", "tab_count", "csv_filenames", "per_tab"):
            assert key in result
        assert result["format"] == "csv_zip"

    def test_nonexistent_file_never_throws(self, tmp_path):
        result = validate_csv_zip(tmp_path / "nope.zip")
        assert result["valid"] is False
