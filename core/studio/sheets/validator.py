"""Post-export validation for XLSX, CSV, and CSV-ZIP sheet artifacts."""

import csv
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional


def validate_xlsx(
    path: Path,
    expected_sheet_names: Optional[List[str]] = None,
    expected_formula_cells: Optional[int] = None,
) -> Dict[str, Any]:
    """Validate an exported XLSX file. Never throws on malformed input.

    Returns a structured result with valid, errors, warnings, sheet_count,
    sheet_names, and formula_cell_count.
    """
    result: Dict[str, Any] = {
        "valid": True,
        "format": "xlsx",
        "errors": [],
        "warnings": [],
        "sheet_count": 0,
        "sheet_names": [],
        "formula_cell_count": 0,
        "chart_count": 0,
        "conditional_format_count": 0,
        "styled_header_sheet_count": 0,
        "quality_score": 0,
    }

    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path))
    except Exception as e:
        result["valid"] = False
        result["errors"].append(f"Failed to open XLSX file: {e}")
        return result

    result["sheet_count"] = len(wb.sheetnames)
    result["sheet_names"] = list(wb.sheetnames)

    # Count formula cells
    formula_count = 0
    chart_count = 0
    conditional_format_count = 0
    styled_header_sheet_count = 0
    has_numeric_data = False

    def _header_styled(sheet) -> bool:
        styled = 0
        header_cells = [c for c in sheet[1] if c.value not in (None, "")]
        for cell in header_cells:
            has_fill = (
                cell.fill is not None
                and cell.fill.fill_type is not None
                and cell.fill.fill_type != "none"
            )
            has_bold = bool(cell.font and cell.font.bold)
            if has_fill and has_bold:
                styled += 1
        return styled > 0

    def _count_conditional_rules(sheet) -> int:
        return sum(len(cf.rules) for cf in sheet.conditional_formatting)

    for ws in wb.worksheets:
        chart_count += len(getattr(ws, "_charts", []))
        conditional_format_count += _count_conditional_rules(ws)
        if _header_styled(ws):
            styled_header_sheet_count += 1
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    has_numeric_data = True

    result["formula_cell_count"] = formula_count
    result["chart_count"] = chart_count
    result["conditional_format_count"] = conditional_format_count
    result["styled_header_sheet_count"] = styled_header_sheet_count

    # Check expected sheet names
    if expected_sheet_names is not None:
        for name in expected_sheet_names:
            # Sanitized names may differ slightly, check with tolerance
            if name not in wb.sheetnames:
                # Try truncated match
                truncated = name[:31]
                if truncated not in wb.sheetnames:
                    result["valid"] = False
                    result["errors"].append(f"Expected sheet '{name}' not found")

    # Check expected formula count
    if expected_formula_cells is not None and formula_count < expected_formula_cells:
        result["valid"] = False
        result["errors"].append(
            f"Expected at least {expected_formula_cells} formula cells, found {formula_count}"
        )

    if has_numeric_data and chart_count == 0:
        result["warnings"].append(
            "Workbook contains numeric data but no charts were detected"
        )

    # Quality score (0-100)
    quality = 0.0
    if result["sheet_count"] > 0:
        quality += 35.0

    if expected_formula_cells is None:
        quality += 20.0 if formula_count > 0 else 12.0
    elif expected_formula_cells <= 0:
        quality += 20.0
    else:
        quality += 20.0 * min(formula_count / expected_formula_cells, 1.0)

    if result["sheet_count"] > 0:
        quality += 20.0 * (styled_header_sheet_count / result["sheet_count"])

    if has_numeric_data:
        quality += 15.0 if chart_count > 0 else 0.0
        quality += 10.0 if conditional_format_count > 0 else 0.0
    else:
        quality += 25.0

    result["quality_score"] = max(0, min(100, int(round(quality))))

    wb.close()
    return result


def validate_csv(
    path: Path,
    min_rows: int = 1,
) -> Dict[str, Any]:
    """Validate an exported CSV file. Never throws on malformed input.

    Returns a structured result with valid, errors, warnings, row_count,
    and column_count.
    """
    result: Dict[str, Any] = {
        "valid": True,
        "format": "csv",
        "errors": [],
        "warnings": [],
        "row_count": 0,
        "column_count": 0,
    }

    try:
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
    except Exception as e:
        result["valid"] = False
        result["errors"].append(f"Failed to read CSV file: {e}")
        return result

    if not rows:
        result["valid"] = False
        result["errors"].append("CSV file is empty")
        return result

    result["column_count"] = len(rows[0])
    # Data rows (excluding header)
    data_rows = rows[1:] if len(rows) > 1 else []
    result["row_count"] = len(data_rows)

    if result["row_count"] < min_rows:
        result["valid"] = False
        result["errors"].append(
            f"Expected at least {min_rows} data rows, found {result['row_count']}"
        )

    return result


def validate_csv_zip(
    path: Path,
    expected_tab_names: Optional[List[str]] = None,
    min_rows: int = 0,
) -> Dict[str, Any]:
    """Validate a ZIP archive containing one CSV per tab. Never throws.

    Returns a structured result with valid, format, errors, warnings,
    tab_count, csv_filenames, and per_tab validation details.
    """
    result: Dict[str, Any] = {
        "valid": True,
        "format": "csv_zip",
        "errors": [],
        "warnings": [],
        "tab_count": 0,
        "csv_filenames": [],
        "per_tab": [],
    }

    try:
        if not zipfile.is_zipfile(path):
            result["valid"] = False
            result["errors"].append("File is not a valid ZIP archive")
            return result
    except Exception as e:
        result["valid"] = False
        result["errors"].append(f"Failed to check ZIP file: {e}")
        return result

    try:
        with zipfile.ZipFile(path, "r") as zf:
            csv_members = [n for n in zf.namelist() if n.endswith(".csv")]
            result["tab_count"] = len(csv_members)
            result["csv_filenames"] = csv_members

            if not csv_members:
                result["valid"] = False
                result["errors"].append("ZIP archive contains no .csv files")
                return result

            if expected_tab_names is not None and len(csv_members) != len(expected_tab_names):
                result["valid"] = False
                result["errors"].append(
                    f"Expected {len(expected_tab_names)} CSV files, found {len(csv_members)}"
                )

            # Validate each CSV member
            with tempfile.TemporaryDirectory() as tmpdir:
                for member in csv_members:
                    extracted = Path(tmpdir) / member
                    extracted.parent.mkdir(parents=True, exist_ok=True)
                    with open(extracted, "wb") as f:
                        f.write(zf.read(member))

                    tab_result = validate_csv(extracted, min_rows=min_rows)
                    result["per_tab"].append({
                        "filename": member,
                        **tab_result,
                    })
                    if not tab_result["valid"]:
                        result["valid"] = False
                        for err in tab_result["errors"]:
                            result["errors"].append(f"{member}: {err}")

    except Exception as e:
        result["valid"] = False
        result["errors"].append(f"Failed to read ZIP archive: {e}")

    return result
