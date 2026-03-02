"""XLSX export via openpyxl for sheet artifacts."""

import re
from pathlib import Path

from core.schemas.studio_schema import SheetContentTree

# Characters invalid in Excel worksheet names
_INVALID_CHARS = re.compile(r"[\[\]:*?/\\]")
_MAX_SHEET_NAME_LEN = 31


def sanitize_sheet_name(name: str) -> str:
    """Sanitize a worksheet name for Excel compatibility."""
    cleaned = _INVALID_CHARS.sub("", name)
    return cleaned[:_MAX_SHEET_NAME_LEN] if cleaned else "Sheet"


def export_to_xlsx(content_tree: SheetContentTree, output_path: Path) -> None:
    """Export a SheetContentTree to XLSX format.

    Creates one worksheet per tab with headers, data, formulas, and column widths.
    Creates parent directories if they don't exist.
    Raises ValueError if the content tree has no tabs.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not content_tree.tabs:
        raise ValueError("Cannot export empty sheet (no tabs)")

    # Ensure output directory exists
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    used_names = set()
    for tab in content_tree.tabs:
        name = sanitize_sheet_name(tab.name)
        # Ensure unique worksheet name
        base_name = name
        counter = 2
        while name in used_names:
            suffix = f" ({counter})"
            name = base_name[: _MAX_SHEET_NAME_LEN - len(suffix)] + suffix
            counter += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)

        # Write headers in row 1
        for col_idx, header in enumerate(tab.headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows starting at row 2
        for row_idx, row_data in enumerate(tab.rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Write formulas
        for cell_addr, formula in tab.formulas.items():
            ws[cell_addr] = formula

        # Apply column widths
        if tab.column_widths:
            for col_idx, width in enumerate(tab.column_widths, start=1):
                col_letter = get_column_letter(col_idx)
                # openpyxl width is in characters; convert from our pixel-ish units
                ws.column_dimensions[col_letter].width = max(width / 7, 8)
        else:
            # Auto-fit fallback
            for col_idx, header in enumerate(tab.headers, start=1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(len(str(header)) + 4, 12)

        # Freeze top row
        ws.freeze_panes = "A2"

    wb.save(str(output_path))
